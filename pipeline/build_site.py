#!/usr/bin/env python3
"""
Build the FLASH Wiki deliverables from flash_library.json:
  1) flash_library.xlsx        - filterable master workbook (WG working copy)
  2) flash_wiki_site/          - deployable MkDocs Material site
       - dashboard home, per-category pages, statistics (charts),
         methodology, downloads
       - GitHub Actions auto-deploy, requirements, .gitignore
Run after flash_harvest.py.  Usage: python3 build_site.py
"""
import os, json, re, html, shutil
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, "flash_library.json"), encoding="utf-8"))
RECS = DATA["records"]
GEN = DATA["generated"]
QUERY = DATA.get("query", "")
OA = sum(1 for r in RECS if r["pmc"])

CATEGORY_ORDER = [
    "Radiobiology", "Physics & Dosimetry", "Modeling & Mechanisms",
    "Beam Delivery & Technology", "Treatment Planning & Optimization",
    "Clinical & Translational", "Reviews & Consensus",
    "Opinions & Debate", "Point-Counterpoint",
    "Perspectives & Commentary", "Uncategorized",
]
CAT_DESC = {
    "Radiobiology": "In vitro, in vivo and mechanistic studies of the FLASH effect, normal-tissue sparing, tumor response, oxygen and immune involvement.",
    "Physics & Dosimetry": "Detectors, reference dosimetry, beam monitoring and dose measurement under ultra-high dose-rate conditions.",
    "Modeling & Mechanisms": "Monte Carlo, radiochemistry, oxygen-depletion and kinetic models explaining the FLASH effect.",
    "Beam Delivery & Technology": "Accelerators, LINAC conversions, laser-driven and VHEE sources, and beam-delivery hardware for UHDR.",
    "Treatment Planning & Optimization": "Dose-rate-aware planning, optimization algorithms and delivery strategies for FLASH.",
    "Clinical & Translational": "Clinical trials, veterinary studies, first-in-human experience and translational workflow.",
    "Reviews & Consensus": "Review articles, roadmaps, consensus statements and guidance documents.",
    "Opinions & Debate": "Opinion essays and forward-looking perspectives on the promise, challenges and future of FLASH radiotherapy — argument rather than systematic evidence synthesis.",
    "Point-Counterpoint": "Formal debate-column articles: Medical Physics Point/Counterpoint and the JACMP three-discipline collaborative radiation therapy (3DCRT) special debates, in which two parties argue opposing propositions.",
    "Perspectives & Commentary": "Editorials, letters, comments and retraction notices on FLASH radiotherapy.",
    "Uncategorized": "FLASH-relevant records awaiting manual category assignment by a WG curator.",
}

def slug(c): return re.sub(r"[^a-z0-9]+", "-", c.lower()).strip("-")
def tldr(abstract):
    if not abstract: return ""
    a = re.sub(r"^[A-Z][A-Za-z /]+:\s*", "", abstract.strip())
    sents = re.split(r"(?<=[.!?])\s+", a)
    out = " ".join(sents[:2])
    return out[:400] + ("…" if len(out) > 400 else "")
def esc(s):
    s = html.escape(s or "")
    return s.replace("[", "\\[").replace("]", "\\]")

# ----------------------------- XLSX -----------------------------------------
def build_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E79"); hdr_font = Font(color="FFFFFF", bold=True)
    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"
    cols = ["PMID","Year","Title","Authors","Journal","Category","Tags","TL;DR","DOI","PMC (open access)","PubMed URL"]
    ws = wb.active; ws.title = "Master Library"; ws.append(cols)
    for r in sorted(RECS, key=lambda x: (x["category"], x.get("year") or "0")):
        ws.append([r["pmid"], r["year"], r["title"], "; ".join(r["authors"]), r["journal"],
                   r["category"], "; ".join(r["tags"]), tldr(r["abstract"]), r["doi"], r["pmc"], r["url"]])
    for i, w in enumerate([10,6,60,34,26,26,30,60,24,16,34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, len(cols))
    s = wb.create_sheet("Summary", 0)
    s.append(["FLASH-RT Living Literature Library"]); s.append([f"Generated {GEN}  |  {len(RECS)} records"])
    s.append([]); s.append(["Category", "Count"])
    cc = Counter(r["category"] for r in RECS)
    for c in CATEGORY_ORDER:
        if cc.get(c): s.append([c, cc[c]])
    s["A1"].font = Font(size=14, bold=True); s.column_dimensions["A"].width = 34; s.column_dimensions["B"].width = 10
    for cell in ("A4","B4"): s[cell].fill = hdr_fill; s[cell].font = hdr_font
    wb.save(os.path.join(HERE, "flash_library.xlsx")); print("wrote flash_library.xlsx")

# ----------------------------- charts ---------------------------------------
def build_charts(assets):
    import matplotlib; matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    NAVY = "#1F4E79"; ACC = "#2E74B5"
    os.makedirs(assets, exist_ok=True)
    # pubs per year
    yc = Counter(int(r["year"]) for r in RECS if r["year"].isdigit())
    yrs = sorted(y for y in yc if y >= 2000); vals = [yc[y] for y in yrs]
    fig, ax = plt.subplots(figsize=(8, 3.2), dpi=120)
    ax.bar(yrs, vals, color=ACC)
    ax.set_title("FLASH-RT publications per year", color=NAVY, fontweight="bold")
    ax.set_xlabel("Year"); ax.set_ylabel("Papers"); ax.spines[["top","right"]].set_visible(False)
    fig.tight_layout(); fig.savefig(os.path.join(assets, "per_year.png")); plt.close(fig)
    # by category
    cc = Counter(r["category"] for r in RECS)
    cats = [c for c in CATEGORY_ORDER if cc.get(c)][::-1]; cv = [cc[c] for c in cats]
    fig, ax = plt.subplots(figsize=(8, 3.8), dpi=120)
    ax.barh(cats, cv, color=NAVY)
    for i, v in enumerate(cv): ax.text(v + 3, i, str(v), va="center", fontsize=8)
    ax.set_title("Papers by category", color=NAVY, fontweight="bold")
    ax.spines[["top","right"]].set_visible(False); fig.tight_layout()
    fig.savefig(os.path.join(assets, "by_category.png")); plt.close(fig)
    # top journals
    jc = Counter(r["journal"] for r in RECS if r["journal"]).most_common(12)
    labels = [j[:34] for j, _ in jc][::-1]; jv = [n for _, n in jc][::-1]
    fig, ax = plt.subplots(figsize=(8, 4), dpi=120)
    ax.barh(labels, jv, color=ACC); ax.set_title("Top 12 journals", color=NAVY, fontweight="bold")
    ax.spines[["top","right"]].set_visible(False); fig.tight_layout()
    fig.savefig(os.path.join(assets, "top_journals.png")); plt.close(fig)
    print("wrote 3 charts")

# ----------------------------- MkDocs ---------------------------------------
def paper_block(r):
    authors = ", ".join(r["authors"][:6]) + (" et al." if len(r["authors"]) > 6 else "")
    badges = []
    if r["pmc"]: badges.append('<span class="badge oa">Open access</span>')
    for t in r["tags"][:4]: badges.append(f'<span class="badge tag">{esc(t)}</span>')
    links = [f'[PubMed]({r["url"]})']
    if r["doi"]: links.append(f'[DOI](https://doi.org/{r["doi"]})')
    if r["pmc"]: links.append(f'[Full text (PMC)](https://www.ncbi.nlm.nih.gov/pmc/articles/{r["pmc"]}/)')
    tl = tldr(r["abstract"]); md = [f'### {esc(r["title"])}\n',
        f'*{esc(authors)}* — {esc(r["journal"])} ({r["year"]})  ', "\n" + " ".join(badges) + "\n"]
    if tl: md.append(f'\n**TL;DR.** {esc(tl)}\n')
    if r["abstract"]: md.append(f'\n??? note "Abstract"\n    {esc(r["abstract"])}\n')
    md.append("\n" + " · ".join(links) + "\n"); md.append("\n---\n")
    return "\n".join(md)

def build_mkdocs():
    site = os.path.join(HERE, "flash_wiki_site"); docs = os.path.join(site, "docs")
    if os.path.exists(site): shutil.rmtree(site)
    os.makedirs(docs)
    assets = os.path.join(docs, "assets"); build_charts(assets)
    by_cat = defaultdict(list)
    for r in RECS: by_cat[r["category"]].append(r)
    cc = Counter(r["category"] for r in RECS)

    # ---- index / dashboard ----
    cards = []
    for c in CATEGORY_ORDER:
        if cc.get(c):
            cards.append(f'<a class="cat-card" href="{slug(c)}/">'
                         f'<span class="n">{cc[c]}</span><span class="c">{c}</span>'
                         f'<span class="d">{CAT_DESC.get(c,"")}</span></a>')
    idx = f"""# FLASH Radiotherapy Living Literature Wiki

**AAPM · BESC · FLASH Working Group**

A continuously updated, categorized index of Medline-indexed FLASH radiotherapy
(ultra-high dose-rate) literature — with concise summaries and one-click citations.

<div class="stat-row">
<div class="stat"><b>{len(RECS):,}</b><span>curated papers</span></div>
<div class="stat"><b>{len([c for c in CATEGORY_ORDER if cc.get(c)])}</b><span>categories</span></div>
<div class="stat"><b>{OA:,}</b><span>open access</span></div>
<div class="stat"><b>{GEN}</b><span>last updated</span></div>
</div>

## Browse by category

<div class="cat-grid">
{''.join(cards)}
</div>

![Publications per year](assets/per_year.png)

## Using this Wiki

Use the **search box** (top) to query across every title, abstract and author in the
corpus. Each entry links to PubMed and, where available, DOI and open-access full text.
Every paper shows a one-line **TL;DR** with the full abstract one click away.

This site is generated from a single source of truth (`flash_library.json`) produced by
an automated PubMed harvest pipeline. Re-running the pipeline refreshes every page — see
**[Methodology](methodology.md)** for the full workflow, and **[Downloads](downloads.md)**
for the master spreadsheet and citation library.
"""
    open(os.path.join(docs, "index.md"), "w", encoding="utf-8").write(idx)

    # ---- category pages ----
    nav_cats = []
    for c in CATEGORY_ORDER:
        recs = by_cat.get(c)
        if not recs: continue
        nav_cats.append((c, f"{slug(c)}.md"))
        recs = sorted(recs, key=lambda x: (x.get("year") or "0"), reverse=True)
        page = [f"# {c}\n", f"{CAT_DESC.get(c,'')}\n", f"*{len(recs)} records. Newest first.*\n", "---\n"]
        page += [paper_block(r) for r in recs]
        open(os.path.join(docs, f"{slug(c)}.md"), "w", encoding="utf-8").write("\n".join(page))

    # ---- statistics ----
    stats = f"""# Statistics

Corpus generated **{GEN}** — **{len(RECS):,}** curated FLASH-RT records,
**{OA:,}** with open-access full text.

## Publications per year
![Per year](assets/per_year.png)

## Papers by category
![By category](assets/by_category.png)

## Most active journals
![Top journals](assets/top_journals.png)
"""
    open(os.path.join(docs, "statistics.md"), "w", encoding="utf-8").write(stats)

    # ---- methodology ----
    meth = f"""# Methodology & provenance

## Corpus definition
Every record derives from one explicit, version-controlled PubMed query executed via the
NCBI E-utilities API. The query is tuned for **high recall** with a downstream precision
filter — the appropriate trade-off for a document meant to capture *all* FLASH work.

```
{QUERY}
```

## Relevance screening
The broad final clause deliberately over-captures, pulling in unrelated "flash" papers
(flash memory, flash nanoprecipitation, photochemistry). A relevance gate then requires
each record to contain a radiotherapy-specific signal; those that do not are moved to an
auditable screening list (`flash_screened_out.csv`) rather than deleted, so a curator can
reinstate any wrongly excluded paper.

## Categorization
Records are auto-classified into the categories at left by a transparent, weighted keyword
model over each paper's title, abstract and MeSH terms; reviews and consensus documents are
routed by PubMed publication type. Every paper also carries secondary tags. Auto-assignment
is a first pass — a category editor can override any assignment in the master spreadsheet,
and the correction propagates on the next site rebuild.

## Summaries
Each record shows the authors' own peer-reviewed **abstract** as its authoritative summary,
plus an auto-generated one-line **TL;DR**. This is a deliberate scientific-integrity choice:
machine-writing fresh summaries for 1,600+ papers risks subtle misstatement of results.

## Update cadence
A scheduled monthly harvest adds new PMIDs and refreshes existing records; a curator reviews
the new and newly-screened records (~30 min), then one command rebuilds this site. Each
update is committed to version control, giving a dated, citable history of the corpus.
"""
    open(os.path.join(docs, "methodology.md"), "w", encoding="utf-8").write(meth)

    # ---- downloads ----
    dl = f"""# Downloads

The Wiki is generated from a single machine-readable source of truth. All artifacts below
regenerate together on each refresh.

| File | Description |
|---|---|
| `flash_library.xlsx` | Filterable master workbook — summary sheet + all {len(RECS):,} records with categories, tags, TL;DRs, DOIs and links |
| `flash_library.csv` | Same data as plain CSV |
| `flash_library.json` | The machine-readable source of truth that drives every view |
| `flash_library.ris` | One-click import into **Zotero** / EndNote / Mendeley, pre-tagged by category |
| `flash_screened_out.csv` | The auto-screened "flash" homonyms, for audit |
| `flash_harvest.py` | The reusable harvest pipeline — re-run to refresh the corpus |

> These files live alongside this site in the `FLASH_Wiki_Automation` folder. To publish
> them on the live site, copy them into `docs/downloads/` before building.
"""
    open(os.path.join(docs, "downloads.md"), "w", encoding="utf-8").write(dl)

    # ---- CSS ----
    os.makedirs(os.path.join(docs, "stylesheets"), exist_ok=True)
    open(os.path.join(docs, "stylesheets", "extra.css"), "w").write("""
.badge{display:inline-block;padding:2px 8px;margin:2px;border-radius:10px;font-size:.72rem;font-weight:600}
.badge.oa{background:#2e7d32;color:#fff}.badge.tag{background:#e3edf7;color:#1F4E79}
h3{margin-top:1.4em}
.stat-row{display:flex;gap:18px;flex-wrap:wrap;margin:18px 0}
.stat-row .stat{background:#eaf1f8;border-radius:10px;padding:12px 18px;min-width:120px}
.stat-row .stat b{display:block;font-size:1.5rem;color:#1F4E79;line-height:1.1}
.stat-row .stat span{font-size:.72rem;text-transform:uppercase;letter-spacing:.5px;color:#5b6b7b}
.cat-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:14px;margin:14px 0}
.cat-card{display:block;border:1px solid #d8e2ee;border-radius:12px;padding:16px;text-decoration:none;
color:inherit;background:#fff;transition:.15s}
.cat-card:hover{border-color:#2E74B5;box-shadow:0 3px 10px rgba(31,78,121,.12)}
.cat-card .n{font-size:1.6rem;font-weight:700;color:#1F4E79}
.cat-card .c{display:block;font-weight:700;margin:2px 0 6px;color:#183a5c}
.cat-card .d{display:block;font-size:.8rem;color:#5b6b7b;line-height:1.4}
""")

    # ---- mkdocs.yml ----
    nav = "\n".join([f"    - {c}: {p}" for c, p in nav_cats])
    yml = f"""site_name: FLASH Radiotherapy Living Wiki
site_description: AAPM BESC FLASH Working Group - categorized FLASH-RT literature
theme:
  name: material
  palette:
    - scheme: default
      primary: indigo
      accent: indigo
      toggle: {{icon: material/weather-night, name: Dark mode}}
    - scheme: slate
      primary: indigo
      accent: indigo
      toggle: {{icon: material/weather-sunny, name: Light mode}}
  features:
    - search.suggest
    - search.highlight
    - navigation.top
    - navigation.instant
    - navigation.tracking
    - toc.follow
    - content.code.copy
extra_css:
  - stylesheets/extra.css
plugins:
  - search
markdown_extensions:
  - admonition
  - pymdownx.details
  - pymdownx.superfences
  - tables
  - attr_list
nav:
  - Home: index.md
  - Categories:
{nav}
  - Statistics: statistics.md
  - Methodology: methodology.md
  - Downloads: downloads.md
"""
    open(os.path.join(site, "mkdocs.yml"), "w", encoding="utf-8").write(yml)

    # ---- repo scaffolding ----
    open(os.path.join(site, "requirements.txt"), "w").write("mkdocs-material>=9.5\n")
    open(os.path.join(site, ".gitignore"), "w").write("site/\n__pycache__/\n*.pyc\n.DS_Store\n")
    gh = os.path.join(site, ".github", "workflows"); os.makedirs(gh, exist_ok=True)
    open(os.path.join(gh, "deploy.yml"), "w").write("""name: Deploy FLASH Wiki
on:
  push:
    branches: [main]
  workflow_dispatch:
permissions:
  contents: write
jobs:
  deploy:
    runs-on: ubuntu-latest
    steps:
      - uses: actions/checkout@v4
      - uses: actions/setup-python@v5
        with: {python-version: '3.x'}
      - run: pip install -r requirements.txt
      - run: mkdocs gh-deploy --force
""")
    open(os.path.join(site, "README.md"), "w").write(f"""# FLASH Radiotherapy Living Wiki

Generated {GEN} · {len(RECS)} curated FLASH-RT records.

## Preview locally
```bash
pip install -r requirements.txt
mkdocs serve            # http://127.0.0.1:8000
```

## Publish to GitHub Pages
1. Create a GitHub repo and push this folder.
2. The included Action (`.github/workflows/deploy.yml`) auto-deploys on every push to `main`.
   (Or run `mkdocs gh-deploy` once manually.)
3. In repo Settings → Pages, set the source to the `gh-pages` branch.

## Refresh the corpus
Re-run `../flash_harvest.py` then `../build_site.py`, commit, and push — the site redeploys.
""")
    print(f"wrote MkDocs site with {len(nav_cats)} category pages + 4 sections -> flash_wiki_site/")

if __name__ == "__main__":
    build_xlsx()
    build_mkdocs()
